import shutil
import openpyxl as xl
import os 
from copy import copy
from openpyxl.worksheet.datavalidation import DataValidation
import time
#import openpyxl as x1
from PIL import Image
from openpyxl import load_workbook
from openpyxl import Workbook



class Copy_report:
    def copy_report(DUTtype):
        report = 'report.xlsx'
        file = 'D:/python/230209/xls/Templates.xlsx'
        shutil.copyfile(file,report)  

                                  
    def main():
        report = 'report.xlsx'
        file = 'D:/python/230209/xls/Templates.xlsx'
        shutil.copyfile(file,report)  
        # kevin add -s
        wb = xl.load_workbook('report.xlsx')
        #ws = wb['test']
        ws = wb.worksheets[0]
        ws.cell(row = 1 ,column = 1).value = 'kkkkkkk'


        #kevin add -s
        #wb2 = xl.load_workbook('report.xlsx')
        #ws2 = wb2.active
        img2 = Image.open('D:/python/230209/S2.png')
        img2 = img2.resize((437,372))
        img2.save('S2_resize.png')
        img2 = xl.drawing.image.Image('S2_resize.png')
        ws.add_image(img2,'A343')
        
        #
        
        #kevin add -E


        wb.save('report.xlsx')
        path = 'D:/python/230209/summary_1.txt'
        f = open(path, 'r')
        print(f.read())
        f.close()

        with open(path,'r') as f:
           lines = f.readlines()
        print(lines)



        #f = open(path, "a")
        #f.write( "123\n")
        #f.close()

        # 使用r+
        #f = open(path, "r+")
        ##f.close()

        # 使用a+
        #f = open(path, "a+")
        #f.write( "789\n")
        #f.close()




        # kevin add -E

    
Copy_report.main()

